"""
Surgery-level temporal reconstruction with a non-monotonic hidden semi-Markov
model (HSMM).

Reads per-frame class-probability CSVs from a configurable directory. The CSVs
are exported by the frame classifier (in this study, the Inception-ResNet-v2
networks fine-tuned in PyTorch); this temporal stage is classifier-agnostic and
only needs columns `path` and `prob_<ClassName>` for each of the six scene classes.

Method (manuscript Sec. 2.8):
  1. Smooth the per-frame posteriors with a sliding window (Eq. 1).
  2. Learn a data-driven phase->phase transition matrix from the expert phase
     sequences of the TRAINING cases only (per fold; Eq. 3). Self-transitions are
     removed and never-observed transitions receive a small floor epsilon, so the
     model can revisit an earlier phase but is discouraged from doing so unless the
     evidence is strong. No fixed left-to-right order is imposed -> non-monotonic.
  3. Decode with a segmental dynamic program (semi-Markov): each segment scores its
     smoothed log-emission plus the log-transition into it, minus a constant switch
     penalty lambda that suppresses over-segmentation, subject to a per-phase minimum
     duration D_k (Eqs. 2, 4). The first segment is fixed to Pre-Craniotomy.
  4. Read the phase-onset time as the first frame decoded to each phase
     (first-entry semantics), and score it against the expert onset via %MAE.

Boundary estimation runs on a few hundred subsampled candidate positions and is
lightweight (the heavy compute is the frame classifier, not this stage).
"""
import os
import re
import glob
import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

import config

# Class ordering helpers.
PHASE_INDEX = {p: i for i, p in enumerate(config.SCENE_CLASSES)}


# ── Phase time parsing ─────────────────────────────────────────────────

def parse_phase_times(xlsx_path=config.PHASE_TIME_XLSX):
    """Read expert phase-onset annotations.

    Expected columns (first sheet): [case-video id "CC-VV", start time, phase name].
    Phase names must match config.SCENE_CLASSES. Returns
    {case_num: [(video_num, seconds, phase_name), ...]} sorted by (video, time).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    phase_times = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] is None or len(row) < 3 or row[2] is None:
            continue
        parts = str(row[0]).split("-")
        try:
            case_num = int(parts[0])
            video_num = int(parts[1])
        except (ValueError, IndexError):
            continue
        t = row[1]
        if isinstance(t, datetime.time):
            seconds = t.hour * 3600 + t.minute * 60 + t.second
        elif isinstance(t, datetime.timedelta):
            seconds = int(t.total_seconds())
        else:
            seconds = int(t)
        phase_name = str(row[2]).strip()
        phase_times.setdefault(case_num, []).append((video_num, seconds, phase_name))
    for case_num in phase_times:
        phase_times[case_num].sort(key=lambda x: (x[0], x[1]))
    return phase_times


# The driver sets this to the directory holding predictions_fold*.csv so that
# split-video lengths can be derived directly from the frame filenames in the
# prediction CSVs (no access to the raw frame folders / patient data required).
_DURATION_SOURCE_DIR = None


def set_duration_source(pred_dir):
    """Point video-duration discovery at a directory of predictions_fold*.csv."""
    global _DURATION_SOURCE_DIR
    _DURATION_SOURCE_DIR = str(pred_dir)


def get_video_durations(case_num):
    """Per-split-video length (seconds) for a case, from prediction-CSV frame names.

    Each frame filename is ``CaseXXYY_HHMMSS.jpg`` (case XX, split-video YY,
    in-video time HHMMSS); the duration of split-video YY is the max HHMMSS seen.
    """
    durations = {}
    case_prefix = f"Case{case_num:02d}"
    src = _DURATION_SOURCE_DIR or config.SCENE_PRE_DIR
    for csv_path in sorted(glob.glob(os.path.join(src, "predictions_fold*.csv"))):
        try:
            paths = pd.read_csv(csv_path, usecols=["path"])["path"]
        except (ValueError, FileNotFoundError):
            continue
        for p in paths:
            fname = os.path.basename(str(p))
            if not fname.startswith(case_prefix):
                continue
            m = re.match(r"Case(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})\.jpg", fname)
            if m:
                vid_num = int(m.group(2))
                h, mi, s = int(m.group(3)), int(m.group(4)), int(m.group(5))
                t_sec = h * 3600 + mi * 60 + s
                if vid_num not in durations or t_sec > durations[vid_num]:
                    durations[vid_num] = t_sec
    return durations


def build_surgery_timeline(case_num, phase_times_entry):
    """Concatenate split videos into one surgery timeline.

    Returns (video_offsets, expert_transitions) where expert_transitions is a list
    of (global_time_seconds, phase_name) for every annotated onset.
    """
    durations = get_video_durations(case_num)
    if not durations:
        return None, None
    video_nums = sorted(durations.keys())
    video_offsets = {}
    offset = 0
    for vn in video_nums:
        video_offsets[vn] = offset
        offset += durations[vn] + 1
    expert_transitions = []
    for vid_num, time_sec, phase_name in phase_times_entry:
        if vid_num in video_offsets:
            expert_transitions.append((video_offsets[vid_num] + time_sec, phase_name))
    return video_offsets, expert_transitions


def expert_first_starts(expert_transitions):
    """First (earliest) global onset per phase name -> {phase_name: global_time}."""
    starts = {}
    for global_time, phase_name in sorted(expert_transitions, key=lambda x: x[0]):
        if phase_name in PHASE_INDEX and phase_name not in starts:
            starts[phase_name] = global_time
    return starts


# ── Prediction timeline reconstruction ─────────────────────────────────

_FILENAME_RE = re.compile(r"Case(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})\.jpg")


def parse_filename(fname):
    basename = os.path.basename(fname)
    m = _FILENAME_RE.match(basename)
    if not m:
        return None, None, None
    case_num = int(m.group(1))
    vid_num = int(m.group(2))
    h, mi, s = int(m.group(3)), int(m.group(4)), int(m.group(5))
    return case_num, vid_num, h * 3600 + mi * 60 + s


def build_prediction_timeline(pred_csv_path, target_case_num, video_offsets):
    df = pd.read_csv(pred_csv_path)
    prob_cols = [f"prob_{c}" for c in config.SCENE_CLASSES]
    basenames = df["path"].apply(os.path.basename)
    pattern = r"Case(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})\.jpg"
    parsed = basenames.str.extract(pattern)
    parsed.columns = ["case", "vid", "h", "m", "s"]
    parsed = parsed.apply(pd.to_numeric, errors="coerce")
    mask = parsed["case"] == target_case_num
    df_case = df[mask].copy()
    p_case = parsed[mask]
    if df_case.empty:
        return None, None
    vid_nums = p_case["vid"].values
    t_secs = p_case["h"].values * 3600 + p_case["m"].values * 60 + p_case["s"].values
    times = []
    valid = []
    for i, (vn, ts) in enumerate(zip(vid_nums, t_secs)):
        if vn in video_offsets:
            times.append(video_offsets[vn] + ts)
            valid.append(i)
    if not times:
        return None, None
    times = np.array(times)
    probs = df_case.iloc[valid][prob_cols].values
    order = np.argsort(times)
    return times[order], probs[order]


# ── Temporal smoothing (Eq. 1) ─────────────────────────────────────────

def smooth_probabilities(times, probs, window=config.SMOOTHING_WINDOW):
    from scipy.ndimage import uniform_filter1d
    dt = np.median(np.diff(times)) if len(times) > 1 else 1.0
    kernel = max(1, int(round(window / dt)))
    if kernel % 2 == 0:
        kernel += 1
    smoothed = uniform_filter1d(probs, size=kernel, axis=0, mode="nearest")
    row_sums = smoothed.sum(axis=1, keepdims=True)
    smoothed = smoothed / (row_sums + 1e-8)
    return smoothed


# ── Data-driven transition matrix (Eq. 3) ──────────────────────────────

def learn_transition(train_phase_times, floor=None):
    """Estimate log phase->phase transition probabilities from expert sequences.

    `train_phase_times` maps case_num -> [(video, seconds, phase_name), ...] and MUST
    contain training cases only (call once per fold with the held-in cases) so that no
    test-case ordering leaks into the decoder.

    The full annotated sequence is used (recurrences included), self-transitions are
    dropped, and never-observed transitions get a small floor epsilon. Row-normalized;
    returns a K x K matrix of log-probabilities.
    """
    if floor is None:
        floor = config.TRANSITION_FLOOR
    K = config.NUM_SCENE_CLASSES
    counts = np.zeros((K, K))
    for _cn, seq in train_phase_times.items():
        phases = ["PreCraniotomy"] + [p for (_v, _s, p) in seq if p in PHASE_INDEX]
        for a, b in zip(phases[:-1], phases[1:]):
            if a in PHASE_INDEX and b in PHASE_INDEX and a != b:
                counts[PHASE_INDEX[a], PHASE_INDEX[b]] += 1
    A = np.full((K, K), floor)
    np.fill_diagonal(A, 0.0)
    A = A + counts
    A = A / A.sum(axis=1, keepdims=True)
    return np.log(A + 1e-12)


# ── Non-monotonic HSMM decoder (Eqs. 2, 4) ─────────────────────────────

def hsmm_decode(times, probs, logA,
                min_dur=None, n_cand=None, switch_pen=None):
    """Segmental (semi-Markov) DP decode. Returns a per-frame phase-label array.

    Each segment scores its smoothed log-emission (summed over frames) plus the
    log-transition logA[prev, k] into phase k, minus a constant switch penalty. A
    segment for phase k must last at least min_dur[k] seconds. No fixed order is
    imposed, so an earlier phase may recur. The first segment is forced to be
    Pre-Craniotomy (surgery always begins before craniotomy).
    """
    if min_dur is None:
        min_dur = config.MIN_PHASE_DURATION
    if n_cand is None:
        n_cand = config.BOUNDARY_CANDIDATES
    if switch_pen is None:
        switch_pen = config.SWITCH_PENALTY

    n, K = len(times), probs.shape[1]
    step = max(1, n // n_cand)
    idx = np.r_[np.arange(0, n, step), n - 1]
    idx = np.unique(idx)
    m = len(idx)

    logP = np.log(probs + 1e-8)
    cum = np.zeros((n + 1, K))
    cum[1:] = np.cumsum(logP, axis=0)          # cum[b,k]-cum[a,k] = emission of [a,b)
    t_sub = times[idx]
    md = np.array([min_dur[c] for c in config.SCENE_CLASSES])
    NEG = -1e18

    dp = np.full((m, K), NEG)
    back = np.full((m, K, 2), -1)              # backpointer: (prev_i, prev_phase)

    # First segment [0, idx[j]) is Pre-Craniotomy.
    for j in range(m):
        if t_sub[j] - times[0] >= md[0] - 1:
            dp[j, 0] = cum[idx[j], 0] - cum[0, 0]

    # Segmental recursion.
    for j in range(m):
        for k in range(K):
            best = dp[j, k]
            for i in range(j):
                if t_sub[j] - t_sub[i] < md[k] - 1:      # minimum-duration constraint
                    continue
                seg = cum[idx[j], k] - cum[idx[i], k]    # emission of segment [i,j) as phase k
                col = dp[i] + logA[:, k] + seg - switch_pen
                ki = int(np.argmax(col))
                v = col[ki]
                if v > best:
                    best = v
                    dp[j, k] = v
                    back[j, k] = (i, ki)

    # Backtrack from the best terminal state.
    jl = m - 1
    kl = int(np.argmax(dp[jl]))
    segs = []
    j, k = jl, kl
    while j > 0 and back[j, k, 0] >= 0:
        i, kp = back[j, k]
        segs.append((idx[i], idx[j], k))
        j, k = i, kp
    segs.append((0, idx[j], k))
    segs.reverse()

    lab = np.zeros(n, dtype=int)
    for a, b, ph in segs:
        lab[a:b] = ph
    lab[idx[jl]:] = kl
    return lab


def first_entry_times(times, lab):
    """First global time each scored phase is entered -> {phase_name: time or None}."""
    out = {}
    for phase_name in config.TRANSITION_PHASES:
        k = PHASE_INDEX[phase_name]
        w = np.where(lab == k)[0]
        out[phase_name] = float(times[w[0]]) if len(w) else None
    return out


def estimate_transitions(times, probs, logA, **kw):
    """Convenience: smooth -> decode -> first-entry dict of phase onsets."""
    lab = hsmm_decode(times, probs, logA, **kw)
    return first_entry_times(times, lab)


# ── Error analysis ─────────────────────────────────────────────────────

def compute_timepoint_errors(predicted, expert):
    """Absolute onset error per scored phase.

    `predicted` is {phase_name: global_time or None} (HSMM first-entry output).
    `expert` is the list of (global_time, phase_name) onset annotations. Only phases
    present in BOTH the expert annotation and the prediction are scored, matched by
    phase name (first-entry semantics), so a never-entered phase is skipped rather
    than penalized as a boundary at time 0.
    """
    expert_starts = expert_first_starts(expert) if isinstance(expert, list) else dict(expert)
    errors = {}
    for phase_name, global_time in expert_starts.items():
        if phase_name not in config.TRANSITION_PHASES:
            continue
        pred_time = predicted.get(phase_name)
        if pred_time is None:
            continue
        signed_err = pred_time - global_time
        errors[phase_name] = {
            "signed_error": signed_err,
            "absolute_error": abs(signed_err),
            "predicted": pred_time,
            "expert": global_time,
        }
    return errors


# ── Visualization ──────────────────────────────────────────────────────

def plot_surgery_timeline(times, probs, predicted, expert_transitions, case_num, save_path):
    """predicted may be a {phase: time} dict (HSMM) or a list of onset times."""
    if isinstance(predicted, dict):
        boundaries = [predicted[p] for p in config.TRANSITION_PHASES if predicted.get(p) is not None]
    else:
        boundaries = list(predicted)
    fig, axes = plt.subplots(2, 1, figsize=(16, 8), sharex=True)
    colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b"]

    ax = axes[0]
    for k, cls_name in enumerate(config.SCENE_CLASSES):
        ax.plot(times, probs[:, k], color=colors[k], alpha=0.7, label=cls_name)
    for i, bt in enumerate(boundaries):
        ax.axvline(bt, color="red", linestyle="--", alpha=0.5,
                   label="Predicted" if i == 0 else "")
    for i, (gt, pname) in enumerate(expert_transitions):
        ax.axvline(gt, color="black", linestyle="-", alpha=0.5,
                   label="Expert" if i == 0 else "")
    ax.set_ylabel("Probability")
    ax.set_title(f"Case {case_num} - Surgery-level Scene Probabilities")
    ax.legend(loc="upper right", fontsize=8)
    ax.set_ylim(0, 1)

    ax2 = axes[1]
    all_bounds = [times[0]] + sorted(boundaries) + [times[-1]]
    for k in range(min(len(all_bounds) - 1, len(colors))):
        ax2.axvspan(all_bounds[k], all_bounds[k + 1], color=colors[k], alpha=0.6)
    for i, (gt, pname) in enumerate(expert_transitions):
        ax2.axvline(gt, color="black", linewidth=2,
                    label="Expert" if i == 0 else "")
    ax2.set_xlabel("Surgery time (seconds)")
    ax2.set_ylabel("Phase")
    ax2.set_yticks([])
    ax2.set_title("Phase Segmentation (first-entry order)")

    plt.tight_layout()
    plt.savefig(save_path, dpi=150)
    plt.close()


# ── Main pipeline ──────────────────────────────────────────────────────

def run_temporal_analysis(pred_csv_dir, save_dir):
    """Run the non-monotonic HSMM over all folds with per-fold (train-only) transitions."""
    os.makedirs(save_dir, exist_ok=True)
    set_duration_source(pred_csv_dir)
    phase_times = parse_phase_times()
    all_errors = []

    for fold in range(1, 6):
        csv_path = os.path.join(pred_csv_dir, f"predictions_fold{fold}.csv")
        if not os.path.exists(csv_path):
            print(f"[WARN] {csv_path} not found, skipping fold {fold}")
            continue

        # Transition matrix from TRAINING cases only (all cases not tested in this fold).
        test_cases = config.FOLD_TEST_CASES[fold]
        test_nums = {int(c.replace("Case", "")) for c in test_cases}
        train_phase_times = {cn: seq for cn, seq in phase_times.items() if cn not in test_nums}
        logA = learn_transition(train_phase_times)

        for case_str in test_cases:
            case_num = int(case_str.replace("Case", ""))
            if case_num not in phase_times:
                print(f"  [WARN] No phase times for Case{case_num}, skipping")
                continue

            video_offsets, expert_transitions = build_surgery_timeline(
                case_num, phase_times[case_num])
            if video_offsets is None:
                continue

            times, probs = build_prediction_timeline(csv_path, case_num, video_offsets)
            if times is None:
                print(f"  [WARN] No predictions for Case{case_num} in fold {fold}")
                continue

            smoothed = smooth_probabilities(times, probs)
            predicted = estimate_transitions(times, smoothed, logA)

            errors = compute_timepoint_errors(predicted, expert_transitions)
            for phase_name, err in errors.items():
                all_errors.append({"fold": fold, "case": case_num, "phase": phase_name, **err})

            plot_surgery_timeline(
                times, smoothed, predicted, expert_transitions, case_num,
                os.path.join(save_dir, f"timeline_case{case_num:02d}_fold{fold}.png"))
            print(f"  Fold {fold}, Case {case_num}: onsets estimated for "
                  f"{sum(v is not None for v in predicted.values())} phases")

    if all_errors:
        err_df = pd.DataFrame(all_errors)
        err_df.to_csv(os.path.join(save_dir, "timepoint_errors.csv"), index=False)
        summary = err_df.groupby("phase").agg(
            mean_signed=("signed_error", "mean"),
            std_signed=("signed_error", "std"),
            mean_absolute=("absolute_error", "mean"),
            std_absolute=("absolute_error", "std"),
            median_absolute=("absolute_error", "median"),
            n=("signed_error", "count"),
        ).round(1)
        summary.to_csv(os.path.join(save_dir, "timepoint_error_summary.csv"))
        print(f"\nTime-point error summary:\n{summary}")
    return all_errors


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--pred_dir", required=True)
    parser.add_argument("--save_dir", required=True)
    args = parser.parse_args()
    run_temporal_analysis(args.pred_dir, args.save_dir)
